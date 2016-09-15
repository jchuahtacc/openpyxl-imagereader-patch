# coding: utf-8
from __future__ import absolute_import

import os.path
from io import BytesIO

from openpyxl.xml.constants import (PACKAGE_WORKSHEET_RELS,
                                    REL_NS,
                                    PACKAGE_RELS,
                                    PACKAGE_IMAGES,
                                    PACKAGE_DRAWINGS,
                                    DRAWING_NS,
                                    SHEET_DRAWING_NS)
from openpyxl.xml.functions import fromstring
from openpyxl.drawing import Image
from openpyxl.cell import get_column_letter


IMAGE_NS = REL_NS + '/image'
_DRAWING_NS = REL_NS + '/drawing'
PACKAGE_DRAWINGS_RELS = PACKAGE_DRAWINGS + '/' + PACKAGE_RELS


def read_image_file(root, rid, valid_files):
    for node in root:
        if node.attrib['Type'] == IMAGE_NS and node.attrib['Id'] == rid:
            image_file = os.path.split(node.attrib['Target'])[-1]
            image_file = PACKAGE_IMAGES + '/' + image_file

            if image_file in valid_files:
                return image_file

    return None


def read_drawings(ws, drawings_path, archive, valid_files):
    """ Given a worksheet and the XML of its drawings file, links drawings to cells
    """
    drawings_codename = os.path.split(drawings_path)[-1]
    rels_file = PACKAGE_DRAWINGS_RELS + '/' + drawings_codename + '.rels'

    if rels_file not in valid_files:
        return None

    rels_source = archive.read(rels_file)
    rels_root = fromstring(rels_source)

    root = fromstring(archive.read(drawings_path))
    for node in root:
        col, row = 0, 0
        name = u''

        cell_from = node.find('{%s}from' % SHEET_DRAWING_NS)

        if cell_from is not None:
            col = cell_from.find('{%s}col' % SHEET_DRAWING_NS)
            if col is not None:
                col = int(col.text)

            row = cell_from.find('{%s}row' % SHEET_DRAWING_NS)
            if row is not None:
                row = int(row.text)

        cell = ws['%s%s' % (get_column_letter(col + 1), row + 1)]

        pic = node.find('{%s}pic' % SHEET_DRAWING_NS)
        if pic is not None:
            nv_pic_pr = pic.find('{%s}nvPicPr' % SHEET_DRAWING_NS)
            if nv_pic_pr is not None:
                nv_pic_pr = nv_pic_pr.find('{%s}cNvPr' % SHEET_DRAWING_NS)
                if nv_pic_pr is not None:
                    name = nv_pic_pr.attrib.get('name', '')

            blip_fill = pic.find('{%s}blipFill' % SHEET_DRAWING_NS)
            if blip_fill is not None:
                blip = blip_fill.find('{%s}blip' % DRAWING_NS)
                if blip is not None:
                    rid = blip.attrib.get('{%s}embed' % REL_NS)
                    if rid is not None:
                        image_file = read_image_file(rels_root, rid, valid_files)
                        if image_file:
                            img = Image(BytesIO(archive.read(image_file)))
                            img.drawing.name = name
                            img.anchor(cell, anchortype='oneCell')

                            ws.add_image(img)


def get_drawings_file(worksheet_path, archive, valid_files):
    """ Returns the XML filename in the archive which contains the drawings for
        the spreadsheet with codename sheet_codename. Returns None if there is no
        such file
    """
    sheet_codename = os.path.split(worksheet_path)[-1]
    rels_file = PACKAGE_WORKSHEET_RELS + '/' + sheet_codename + '.rels'

    if rels_file not in valid_files:
        return None

    rels_source = archive.read(rels_file)
    root = fromstring(rels_source)
    for node in root:
        if node.attrib['Type'] == _DRAWING_NS:
            drawings_file = os.path.split(node.attrib['Target'])[-1]
            drawings_file = PACKAGE_DRAWINGS + '/' + drawings_file

            if drawings_file in valid_files:
                return drawings_file

    return None
