from openpyxl.reader import excel

from patch_reader.excel import _load_workbook


setattr(excel, '_load_workbook', _load_workbook)
